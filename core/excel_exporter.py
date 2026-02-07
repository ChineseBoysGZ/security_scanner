"""
core/excel_exporter.py - 修正版（修复Excel写入顺序）
Excel导出器 - 支持流式写入、分批写入、性能优化
修复write_only模式下的表头写入和列宽设置问题
"""
import os
import warnings
import threading
import time
from typing import List, Optional, Dict, Any, Iterator
from pathlib import Path
from queue import Queue, Empty
from datetime import datetime
import logging

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

from config.constants import EXCEL_MAX_CELL_LEN, DEFAULT_EXCEL_HEADERS, DEFAULT_EXCEL_WIDTHS, DEFAULT_EXCEL_EXT
from core.scan_models import ScanResult
from utils.log_utils import logger

# 忽略openpyxl的警告
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


class ExcelStreamWriter:
    """
    Excel流式写入器 - 异步写入，避免阻塞主线程
    支持实时追加结果，内存高效。
    """
    
    def __init__(self, file_path: str, headers: List[str], column_widths: List[int]):
        self.file_path = file_path
        self.headers = headers
        self.column_widths = column_widths
        
        # 写入队列
        self.write_queue = Queue(maxsize=5000)
        self.stop_event = threading.Event()
        self.write_thread: Optional[threading.Thread] = None
        
        # 统计信息
        self.rows_written = 0
        self.start_time = time.time()
        
        # 初始化工作簿（⚠️ 修复：不使用write_only模式，因为它不支持列宽设置）
        self._init_workbook()
    
    def _init_workbook(self):
        """初始化工作簿（⚠️ 修复：使用普通模式确保列宽设置生效）"""
        # 创建普通工作簿
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "扫描结果"
        
        # ✅ 修复：首先写入表头
        self.worksheet.append(self.headers)
        self.rows_written = 1  # 表头占一行
        
        # ✅ 修复：设置列宽（必须在有数据后才能设置）
        for col_idx, width in enumerate(self.column_widths, 1):
            column_letter = get_column_letter(col_idx)
            # 确保列宽在合理范围内
            actual_width = min(max(width, 5), 50)  # 限制在5-50之间
            self.worksheet.column_dimensions[column_letter].width = actual_width
        
        # 设置表头样式
        self._apply_header_style()
        
        logger.info(f"Excel工作簿初始化完成，表头: {self.headers}")
    
    def _apply_header_style(self):
        """应用表头样式"""
        import platform
        
        # 确定字体
        system = platform.system()
        if system == 'Darwin':  # Mac
            font_name = 'PingFang SC'
        elif system == 'Windows':  # Windows
            font_name = 'Arial Unicode MS'
        else:  # Linux/其他
            font_name = 'Arial'
        
        # 创建表头字体样式
        header_font = Font(name=font_name, bold=True, size=12)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 应用样式到表头行
        for cell in self.worksheet[1]:
            cell.font = header_font
            cell.alignment = header_alignment
    
    def _clean_cell_content(self, content: str) -> str:
        """清洗单元格内容，移除非法字符并截断"""
        if not content:
            return ""
        
        if not isinstance(content, str):
            content = str(content)
        
        # 保留：可打印ASCII、中文字符、换行符、制表符
        valid_chars = []
        for char in content:
            code = ord(char)
            if (32 <= code <= 126) or (0x4e00 <= code <= 0x9fa5) or char in ('\n', '\t'):
                valid_chars.append(char)
        
        cleaned = ''.join(valid_chars)
        
        # 截断过长的内容
        if len(cleaned) > EXCEL_MAX_CELL_LEN:
            cleaned = cleaned[:EXCEL_MAX_CELL_LEN] + "【内容超长已截断】"
        
        return cleaned
    
    def _scanresult_to_row(self, result: ScanResult) -> List:
        """将ScanResult转换为Excel行数据"""
        row = [None] * len(self.headers)
        
        # 核心字段映射
        for i, header in enumerate(self.headers):
            header_lower = header.lower()
        
            if '规则' in header or 'rule' in header_lower:
                row[i] = result.rule_name
            elif '路径' in header or 'path' in header_lower:
                row[i] = result.file_path
            elif '内容' in header or 'content' in header_lower or '匹配' in header:
                row[i] = self._clean_cell_content(result.match_content)
            elif '应用' in header or 'app' in header_lower:
                row[i] = result.app_id
            elif '系统' in header or 'system' in header_lower:
                row[i] = result.system_name  # ✅ 新增字段
            elif '负责' in header or 'person' in header_lower:
                row[i] = result.responsible_person  # ✅ 新增字段
            elif '备注' in header or 'notes' in header_lower or 'remark' in header_lower:
                row[i] = result.notes  # ✅ 新增字段
            elif '最终' in header or 'final' in header_lower or 'result' in header_lower:
                row[i] = result.final_result  # ✅ 新增字段
            elif '行号' in header or 'line' in header_lower:
                row[i] = result.line_num
            elif '严重' in header or 'severity' in header_lower:
                row[i] = result.severity
            elif '大小' in header or 'size' in header_lower:
                row[i] = f"{result.file_size // 1024 if result.file_size else 0}KB"
            elif '时间' in header or 'time' in header_lower:
                row[i] = datetime.fromtimestamp(result.timestamp).strftime('%Y-%m-%d %H:%M:%S')
    
        return row
    
    def start(self):
        """启动写入线程"""
        if self.write_thread and self.write_thread.is_alive():
            return
        
        self.stop_event.clear()
        self.write_thread = threading.Thread(
            target=self._write_worker,
            daemon=True,
            name="ExcelWriterThread"
        )
        self.write_thread.start()
        
        logger.info(f"Excel写入线程已启动，目标文件: {self.file_path}")
    
    def _write_worker(self):
        """写入工作线程"""
        try:
            batch = []
            batch_size = 100  # 每批写入100行
            
            while not self.stop_event.is_set() or not self.write_queue.empty():
                try:
                    # 从队列获取结果（非阻塞）
                    result = self.write_queue.get(timeout=0.5)
                    
                    # 转换为行数据
                    row = self._scanresult_to_row(result)
                    batch.append(row)
                    
                    # 达到批次大小时写入
                    if len(batch) >= batch_size:
                        self._write_batch(batch)
                        batch = []
                    
                    self.write_queue.task_done()
                    
                except Empty:
                    # 队列为空，写入剩余批次
                    if batch:
                        self._write_batch(batch)
                        batch = []
                    continue
                except Exception as e:
                    logger.error(f"Excel写入错误: {e}")
                    continue
            
            # 写入剩余批次
            if batch:
                self._write_batch(batch)
            
            # 保存文件
            self._save_workbook()
            
        except Exception as e:
            logger.error(f"Excel写入线程异常: {e}")
    
    def _write_batch(self, batch: List[List]):
        """批量写入行数据（⚠️ 修复：确保数据从第2行开始）"""
        if not batch:
            return
        
        for row in batch:
            self.worksheet.append(row)
            self.rows_written += 1
        
        # 每写入1000行记录一次日志
        if self.rows_written % 1000 == 0:
            elapsed = time.time() - self.start_time
            logger.info(f"Excel已写入 {self.rows_written} 行（包含表头），耗时: {elapsed:.1f}秒")
            
            # 实时保存（防止数据丢失）
            try:
                self.workbook.save(self.file_path)
                logger.debug(f"Excel已实时保存到: {self.file_path}")
            except Exception as e:
                logger.warning(f"Excel实时保存失败: {e}")
    
    def _save_workbook(self):
        """保存工作簿"""
        try:
            # 应用数据行样式
            self._apply_data_style()
            
            # 保存文件
            self.workbook.save(self.file_path)
            elapsed = time.time() - self.start_time
            
            # 计算数据行数（去掉表头）
            data_rows = self.rows_written - 1
            
            logger.info(
                f"Excel文件保存成功: {self.file_path}\n"
                f"表头行数: 1，数据行数: {data_rows}，总行数: {self.rows_written}\n"
                f"总耗时: {elapsed:.1f}秒，平均速度: {data_rows / max(elapsed, 0.001):.1f} 行/秒"
            )
        except Exception as e:
            logger.error(f"Excel文件保存失败: {e}")
            raise
    
    def _apply_data_style(self):
        """应用数据行样式"""
        import platform
        
        system = platform.system()
        if system == 'Darwin':  # Mac
            font_name = 'PingFang SC'
        elif system == 'Windows':  # Windows
            font_name = 'Arial Unicode MS'
        else:  # Linux/其他
            font_name = 'Arial'
        
        # 数据行字体
        data_font = Font(name=font_name, size=10)
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # 从第2行开始应用样式（第1行是表头）
        for row in self.worksheet.iter_rows(min_row=2, max_row=self.rows_written):
            for cell in row:
                cell.font = data_font
                cell.alignment = data_alignment
    
    def add_result(self, result: ScanResult):
        """添加结果到写入队列"""
        if self.stop_event.is_set():
            return False
        
        try:
            self.write_queue.put(result, timeout=1)
            return True
        except:
            logger.warning("Excel写入队列已满，结果被丢弃")
            return False
    
    def add_results_batch(self, results: List[ScanResult]):
        """批量添加结果到写入队列"""
        success_count = 0
        for result in results:
            if self.add_result(result):
                success_count += 1
        
        return success_count
    
    def stop(self, wait: bool = True):
        """停止写入器"""
        self.stop_event.set()
        
        if wait and self.write_thread:
            # 等待队列处理完成
            self.write_queue.join()
            self.write_thread.join(timeout=10)
            
            if self.write_thread.is_alive():
                logger.warning("Excel写入线程未在10秒内结束")
    
    def get_stats(self) -> Dict[str, Any]:
        """获取统计信息"""
        elapsed = time.time() - self.start_time
        data_rows = max(0, self.rows_written - 1)  # 减去表头行
        
        return {
            'file_path': self.file_path,
            'header_rows': 1,
            'data_rows': data_rows,
            'total_rows': self.rows_written,
            'queue_size': self.write_queue.qsize(),
            'elapsed_seconds': round(elapsed, 2),
            'rows_per_second': round(data_rows / max(elapsed, 0.001), 1)
        }


class ExcelExporter:
    """
    Excel导出器 - 高级API
    提供多种导出模式和性能优化。
    """
    
    @staticmethod
    def export(
        results: List[ScanResult],
        save_path: str,
        headers: List[str],
        column_widths: List[int],
        streaming: bool = True
    ) -> bool:
        """
        导出扫描结果到Excel
        
        :param results: 扫描结果列表
        :param save_path: 保存路径
        :param headers: 表头列表
        :param column_widths: 列宽列表
        :param streaming: 是否使用流式写入（推荐True）
        :return: 是否成功
        """
        if not results:
            logger.warning("无扫描结果可导出")
            return False
        
        # 确保保存路径有正确的扩展名
        if not save_path.lower().endswith('.xlsx'):
            save_path += '.xlsx'
        
        if streaming:
            return ExcelExporter._export_streaming(results, save_path, headers, column_widths)
        else:
            return ExcelExporter._export_batch(results, save_path, headers, column_widths)
    
    @staticmethod
    def _export_streaming(
        results: List[ScanResult],
        save_path: str,
        headers: List[str],
        column_widths: List[int]
    ) -> bool:
        """流式导出（内存高效）"""
        try:
            # 创建写入器
            writer = ExcelStreamWriter(save_path, headers, column_widths)
            writer.start()
            
            # 分批写入结果（避免一次性占用太多内存）
            batch_size = 1000
            total_count = len(results)
            
            for i in range(0, total_count, batch_size):
                batch = results[i:i + batch_size]
                success_count = writer.add_results_batch(batch)
                
                # 记录进度
                if i % 5000 == 0:
                    processed = min(i + batch_size, total_count)
                    logger.info(f"导出进度: {processed}/{total_count} ({processed/total_count*100:.1f}%)")
            
            # 停止写入器并等待完成
            writer.stop(wait=True)
            
            # 验证导出结果
            stats = writer.get_stats()
            if stats['data_rows'] == total_count:
                logger.info(f"✅ Excel导出验证通过：预期{total_count}行，实际导出{stats['data_rows']}行")
            else:
                logger.warning(f"⚠️ Excel导出数量不匹配：预期{total_count}行，实际导出{stats['data_rows']}行")
            
            return True
            
        except Exception as e:
            logger.error(f"流式导出失败: {e}")
            return False
    
    @staticmethod
    def _export_batch(
        results: List[ScanResult],
        save_path: str,
        headers: List[str],
        column_widths: List[int]
    ) -> bool:
        """批量导出（传统方式，兼容性）"""
        try:
            # 创建传统工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "扫描结果"
            
            # ✅ 修复：写入表头
            ws.append(headers)
            
            # ✅ 修复：设置列宽（在写入数据后）
            for col, width in enumerate(column_widths, 1):
                column_letter = get_column_letter(col)
                ws.column_dimensions[column_letter].width = width
            
            # 写入数据
            count = 0
            for result in results:
                row = ExcelExporter._result_to_row(result, headers)
                ws.append(row)
                count += 1
                
                # 进度提示
                if count % 1000 == 0:
                    logger.debug(f"已写入 {count} 条记录")
            
            # 保存文件
            wb.save(save_path)
            
            logger.info(f"Excel导出成功: {save_path}，共 {count} 条记录")
            return True
            
        except Exception as e:
            logger.error(f"批量导出失败: {e}")
            return False
    
    @staticmethod
    def _result_to_row(result: ScanResult, headers: List[str]) -> List:
        """转换结果为行数据"""
        row = [None] * len(headers)
        
        for i, header in enumerate(headers):
            header_lower = header.lower()
            
            if '规则' in header or 'rule' in header_lower:
                row[i] = result.rule_name
            elif '路径' in header or 'path' in header_lower:
                row[i] = result.file_path
            elif '内容' in header or 'content' in header_lower or '匹配' in header:
                # 清洗内容
                content = result.match_content
                if len(content) > EXCEL_MAX_CELL_LEN:
                    content = content[:EXCEL_MAX_CELL_LEN] + "【内容超长已截断】"
                row[i] = content
            elif '应用' in header or 'app' in header_lower:
                row[i] = result.app_id
            elif '行号' in header or 'line' in header_lower:
                row[i] = result.line_num
            elif '严重' in header or 'severity' in header_lower:
                row[i] = result.severity
        
        return row
    
    @staticmethod
    def create_writer(file_path: str, headers: List[str], column_widths: List[int]) -> ExcelStreamWriter:
        """
        创建流式写入器（用于实时导出）
        
        :return: ExcelStreamWriter实例
        """
        return ExcelStreamWriter(file_path, headers, column_widths)


# 兼容原有函数
def export_excel(
    results: List[ScanResult],
    save_path: str,
    headers: List[str],
    column_widths: List[int]
) -> bool:
    """
    兼容原有接口的导出函数
    默认使用流式写入以提高性能
    """
    return ExcelExporter.export(results, save_path, headers, column_widths, streaming=True)


def clean_excel_invalid_chars(s: str) -> str:
    """
    过滤openpyxl不支持的不可打印字符（兼容原有函数）
    
    注意：新的写入器会自动处理字符清洗，此函数主要用于兼容性
    """
    if not isinstance(s, str):
        return str(s) if s else ""
    
    valid_chars = []
    for char in s:
        code = ord(char)
        # 保留：可打印ASCII、中文字符、换行符、制表符
        if (32 <= code <= 126) or (0x4e00 <= code <= 0x9fa5) or char in ('\n', '\t'):
            valid_chars.append(char)
    
    return ''.join(valid_chars)
